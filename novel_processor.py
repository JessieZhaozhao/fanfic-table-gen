"""
小说文档处理脚本
功能：
  1. 读取指定目录下的 .docx 小说文档
  2. 从文件名提取标题
  3. 提取正文内容
  4. 在每篇正文末尾追加提醒语 + 根据角色名自动匹配游戏标签
  5. 输出为格式化的 Excel 表格（A列=小说名, B列=正文）
"""

import os
import re
import sys
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# 配置区 — 在这里修改输入文件和输出路径
# ============================================================

# 输入的 docx 文件列表（支持拖入文件路径）
INPUT_FILES = [
    r"D:/小说/生文合集3/分手后喝醉跑去前男友左然家洗澡.docx",
    r"D:/小说/生文合集3/故意想让夏萧因主动求和那之后就只能好好哄着了呀.docx",
    r"D:/小说/生文合集3/买下娇夫人夏萧因后你学会了当引导型恋人.docx",
    r"D:/小说/生文合集3/夏以昼当皇帝就是为了罔顾规则，谁知妹妹要逃.docx",
    r"D:/小说/生文合集3/以为是自己出轨了，原来是被易遇水煎了.docx",
    r"D:/小说/生文合集3/易遇这种小三上位的防小三最狠了.docx",
]

# 输出 Excel 路径
OUTPUT_XLSX = r"D:/小说/小说正文1_处理后.xlsx"

# ============================================================
# 国乙男主表：游戏 → 角色列表（用于反向匹配）
# ============================================================

GAME_CHARACTER_MAP = {
    "恋与制作人": ["李泽言", "许墨", "白起", "周棋洛"],
    "光与夜之恋": ["陆沉", "齐司礼", "萧逸", "查理苏", "夏鸣星"],
    "未定事件簿": ["左然", "莫弈", "夏彦", "陆景和"],
    "时空中的绘旅人": ["叶瑄", "司岚", "路辰", "罗夏", "艾因"],
    "恋与深空": ["秦彻", "祁煜", "沈星回", "黎深", "夏以昼"],
    "世界之外": ["柏源", "夏萧因", "易遇", "顾时夜"],
}

# 构建反向映射：角色名 → 游戏名（优先匹配较长的名字，避免短名误命中）
CHAR_TO_GAME = {}
for game, characters in GAME_CHARACTER_MAP.items():
    for char in characters:
        CHAR_TO_GAME[char] = game

# 按名字长度降序排列，优先匹配长名字
SORTED_CHARS = sorted(CHAR_TO_GAME.keys(), key=len, reverse=True)

# ============================================================
# 固定追加内容
# ============================================================

REMINDER_TEXT = "提醒一下宝宝们不用蹲，后续4连进群⬇️自取就可以啦~不会在这里更新哦 ᗜ ‸ ᗜ"

FIXED_TAGS = ["#bg", "#乙女向", "#同人文"]


def extract_title(filepath):
    """从文件路径中提取文件名（不含扩展名）作为小说标题"""
    return os.path.splitext(os.path.basename(filepath))[0]


def read_docx_content(filepath):
    """读取 docx 文件的全部正文段落，合并为字符串"""
    doc = Document(filepath)
    paragraphs = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def detect_character_and_game(title):
    """
    根据小说标题中出现的国乙男主角色名，匹配其所属游戏。
    返回 (角色名, 游戏名) 或 (None, None)
    """
    for char in SORTED_CHARS:
        if char in title:
            return char, CHAR_TO_GAME[char]
    return None, None


def append_reminder_and_tags(content, title):
    """
    在正文末尾：
      1. 追加提醒语（换行）
      2. 追加标签（换行）：#角色名 #游戏名 #bg #乙女向 #同人文
    """
    lines = [content]

    # 1. 提醒语
    lines.append("")
    lines.append(REMINDER_TEXT)

    # 2. 匹配角色 & 游戏
    char_name, game_name = detect_character_and_game(title)

    if char_name and game_name:
        tags = f"#{char_name} #{game_name} {' '.join(FIXED_TAGS)}"
    else:
        # 未匹配到角色时，只加通用标签
        tags = " ".join(FIXED_TAGS)

    lines.append("")
    lines.append(tags)

    return "\n".join(lines)


def process_files(input_files, output_path):
    """处理所有文件并生成 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "6.17"

    # ---------- 表头样式 ----------
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # 写表头
    ws["A1"] = "小说名"
    ws["B1"] = "正文"
    for col in ["A", "B"]:
        cell = ws[f"{col}1"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ---------- 数据行 ----------
    content_alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )
    title_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 2
    success_count = 0
    fail_list = []

    for filepath in input_files:
        if not os.path.exists(filepath):
            print(f"[⚠️ 文件不存在] {filepath}")
            fail_list.append((filepath, "文件不存在"))
            continue

        try:
            title = extract_title(filepath)
            print(f"[{row_idx - 1}] 正在处理: {title}")

            # 读取正文
            raw_content = read_docx_content(filepath)
            if not raw_content.strip():
                print(f"  ⚠️ 文档内容为空，跳过")
                fail_list.append((filepath, "文档为空"))
                continue

            # 追加提醒语 + 标签
            final_content = append_reminder_and_tags(raw_content, title)

            # 写入 Excel
            ws.cell(row=row_idx, column=1, value=title)
            ws.cell(row=row_idx, column=2, value=final_content)

            # 设置样式
            title_cell = ws.cell(row=row_idx, column=1)
            content_cell = ws.cell(row=row_idx, column=2)

            title_cell.alignment = title_alignment
            title_cell.border = thin_border
            title_cell.font = Font(size=10)

            content_cell.alignment = content_alignment
            content_cell.border = thin_border
            content_cell.font = Font(size=10)

            # 匹配结果反馈
            char_name, game_name = detect_character_and_game(title)
            if char_name:
                print(f"  ✅ 匹配角色: {char_name} → {game_name}")
            else:
                print(f"  ⚠️ 未匹配到已知角色")

            success_count += 1
            row_idx += 1

        except Exception as e:
            print(f"  ❌ 处理失败: {e}")
            fail_list.append((filepath, str(e)))
            continue

    # ---------- 列宽调整 ----------
    # A列（标题）：宽度 30
    ws.column_dimensions['A'].width = 35
    # B列（正文）：宽度 120
    ws.column_dimensions['B'].width = 120

    # 冻结首行
    ws.freeze_panes = "A2"

    # ---------- 保存 ----------
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    print(f"\n{'='*50}")
    print(f"✅ 处理完成！")
    print(f"   成功: {success_count} 篇")
    print(f"   失败: {len(fail_list)} 篇")
    if fail_list:
        for fp, reason in fail_list:
            print(f"     - {os.path.basename(fp)}: {reason}")
    print(f"📁 输出文件: {output_path}")
    print(f"{'='*50}")


# ============================================================
# 入口
# ============================================================

if __name__ == "__main__":
    process_files(INPUT_FILES, OUTPUT_XLSX)
