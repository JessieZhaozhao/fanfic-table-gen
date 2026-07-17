# -*- coding: utf-8 -*-
"""
CLI六层梗表格生成脚本
从小说正文1_617.xlsx读取数据，结合AI分析结果，生成CLI_六层梗格式的Excel表格
A列：洗后梗（世界观+六层梗分析）
B列：标题
C列：纯正文（去除提醒语、分隔线和标签）
D列：三条选项（文游写手视角，不同XP方向）
"""

import os
import re
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ============================================================
# 配置
# ============================================================
INPUT_XLSX = r"D:/小说/表格/小说正文1_617.xlsx"
OUTPUT_XLSX = r"D:/小说/CLI_六层梗_617.xlsx"
ANALYSIS_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), "novel_analysis_data.json")

# ============================================================
# 国乙男主表
# ============================================================
GAME_CHARACTER_MAP = {
    "恋与制作人": ["李泽言", "许墨", "白起", "周棋洛"],
    "光与夜之恋": ["陆沉", "齐司礼", "萧逸", "查理苏", "夏鸣星"],
    "未定事件簿": ["左然", "莫弈", "夏彦", "陆景和"],
    "时空中的绘旅人": ["叶瑄", "司岚", "路辰", "罗夏", "艾因"],
    "恋与深空": ["秦彻", "祁煜", "沈星回", "黎深", "夏以昼"],
    "世界之外": ["柏源", "夏萧因", "易遇", "顾时夜"],
}
CHAR_TO_GAME = {}
for game, chars in GAME_CHARACTER_MAP.items():
    for c in chars:
        CHAR_TO_GAME[c] = game
SORTED_CHARS = sorted(CHAR_TO_GAME.keys(), key=len, reverse=True)

SEPARATOR_PATTERN = re.compile(r'^-{3,}$')
REMINDER_TEXT = "提醒一下宝宝们不用蹲，后续4连进群⬇️自取就可以啦~不会在这里更新哦 ᗜ ‸ ᗜ"
TAG_PATTERN = re.compile(r'^#[^\s]+(\s+#[^\s]+)+$')


def detect_character_and_game(title):
    for char in SORTED_CHARS:
        if char in title:
            return char, CHAR_TO_GAME[char]
    return None, None


def clean_content(raw_content):
    """去除正文末尾的分隔线、提醒语、标签行"""
    lines = raw_content.split('\n')
    remove_indices = set()
    i = len(lines) - 1

    while i >= 0 and lines[i].strip() == '':
        remove_indices.add(i)
        i -= 1
    if i >= 0 and TAG_PATTERN.match(lines[i].strip()):
        remove_indices.add(i)
        i -= 1
    while i >= 0 and lines[i].strip() == '':
        remove_indices.add(i)
        i -= 1
    if i >= 0 and REMINDER_TEXT in lines[i]:
        remove_indices.add(i)
        i -= 1
    while i >= 0 and lines[i].strip() == '':
        remove_indices.add(i)
        i -= 1
    if i >= 0 and SEPARATOR_PATTERN.match(lines[i].strip()):
        remove_indices.add(i)
        i -= 1
    while i >= 0 and lines[i].strip() == '':
        remove_indices.add(i)
        i -= 1

    cleaned = [lines[j] for j in range(len(lines)) if j not in remove_indices]
    return '\n'.join(cleaned).rstrip('\n')


def load_analysis_data():
    """加载AI分析数据"""
    with open(ANALYSIS_JSON, 'r', encoding='utf-8') as f:
        return json.load(f)


def process_novels():
    analysis_data = load_analysis_data()

    wb_src = load_workbook(INPUT_XLSX)
    ws_src = wb_src.active

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet1"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    content_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    headers = ["洗后梗", "标题", "正文", "选项"]
    for col, h in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_out = 2
    for row_src in range(2, ws_src.max_row + 1):
        title = ws_src.cell(row=row_src, column=1).value
        raw_content = ws_src.cell(row=row_src, column=2).value
        if not title or not raw_content:
            continue

        print(f"[{row_out - 1}] 处理: {title}")

        # Step 1: 清洗正文
        clean_text = clean_content(raw_content)
        print(f"  正文清洗完成 ({len(clean_text)} 字)")

        # Step 2: 匹配角色 & 六层梗
        char_name, game_name = detect_character_and_game(title)
        analysis = analysis_data.get(title, None)

        if analysis:
            world_prefix = "世界观 " + analysis['game'] + "同人"
            six_layers = analysis['six_layers']
            options = analysis['options']
        else:
            world_prefix = "世界观 " + (game_name or "未知") + "同人"
            six_layers = "（需手动填写六层梗分析）"
            options = "（需手动填写选项）"

        a_content = world_prefix + "\n" + six_layers

        # 写入
        cell_a = ws_out.cell(row=row_out, column=1, value=a_content)
        cell_a.alignment = content_alignment
        cell_a.border = thin_border
        cell_a.font = Font(size=10)

        cell_b = ws_out.cell(row=row_out, column=2, value=title)
        cell_b.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_b.border = thin_border
        cell_b.font = Font(size=10)

        cell_c = ws_out.cell(row=row_out, column=3, value=clean_text)
        cell_c.alignment = content_alignment
        cell_c.border = thin_border
        cell_c.font = Font(size=10)

        cell_d = ws_out.cell(row=row_out, column=4, value=options)
        cell_d.alignment = content_alignment
        cell_d.border = thin_border
        cell_d.font = Font(size=10)

        print(f"  完成: {char_name} -> {game_name}")
        row_out += 1

    ws_out.column_dimensions['A'].width = 40
    ws_out.column_dimensions['B'].width = 30
    ws_out.column_dimensions['C'].width = 100
    ws_out.column_dimensions['D'].width = 50
    ws_out.freeze_panes = "A2"

    output_dir = os.path.dirname(OUTPUT_XLSX)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    wb_out.save(OUTPUT_XLSX)
    print(f"\nCLI六层梗表格生成完成！")
    print(f"输出文件: {OUTPUT_XLSX}")


if __name__ == "__main__":
    process_novels()
