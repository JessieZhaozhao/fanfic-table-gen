# -*- coding: utf-8 -*-
"""
终极表格格式化输出脚本
读取 JSON 数据（从网页导出），生成与终极.xlsx样式一致的格式化Excel

用法：
  python format_xlsx.py

输出的xlsx具有：
  - 表头：蓝色背景、加粗、居中
  - 内容：自动换行、顶部对齐
  - 边框：全框线
  - 列宽适配
  - 冻结首行
"""

import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# 配置
# ============================================================
# 输入JSON路径（网页导出的 novels_data.json）
INPUT_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), "novels_data.json")
# 输出xlsx路径
OUTPUT_XLSX = r"D:/小说/表格/终极表格.xlsx"


def load_novels(json_path):
    """从JSON文件加载小说数据"""
    if not os.path.exists(json_path):
        print(f"错误：找不到 {json_path}")
        print("请先从网页导出 JSON 数据文件")
        return []
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if not isinstance(data, list):
        print("错误：JSON 格式不正确，应为数组")
        return []
    return data


def build_xlsx(novels, output_path):
    """生成格式化的xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    content_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    title_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    content_font = Font(size=10)

    # 写表头
    headers = ["洗后梗", "标题", "正文", "引流", "选项"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据
    for i, novel in enumerate(novels):
        row = i + 2
        # A列：洗后梗
        cell_a = ws.cell(row=row, column=1, value=novel.get("columnA", ""))
        cell_a.alignment = content_align
        cell_a.border = thin_border
        cell_a.font = content_font

        # B列：标题
        cell_b = ws.cell(row=row, column=2, value=novel.get("columnB", novel.get("title", "")))
        cell_b.alignment = title_align
        cell_b.border = thin_border
        cell_b.font = content_font

        # C列：正文
        cell_c = ws.cell(row=row, column=3, value=novel.get("columnC", novel.get("content", "")))
        cell_c.alignment = content_align
        cell_c.border = thin_border
        cell_c.font = content_font

        # D列：引流
        cell_d = ws.cell(row=row, column=4, value=novel.get("columnD", ""))
        cell_d.alignment = content_align
        cell_d.border = thin_border
        cell_d.font = content_font

        # E列：选项
        cell_e = ws.cell(row=row, column=5, value=novel.get("columnE", ""))
        cell_e.alignment = content_align
        cell_e.border = thin_border
        cell_e.font = content_font

        print(f"  [{i+1}] {novel.get('title', '未命名')}")

    # 列宽
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 100
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50

    # 冻结首行
    ws.freeze_panes = "A2"

    # 确保输出目录存在
    out_dir = os.path.dirname(output_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output_path)
    print(f"\n格式化xlsx生成完成！")
    print(f"输出文件: {output_path}")


def main():
    print("=== 终极表格格式化输出 ===")
    novels = load_novels(INPUT_JSON)
    if not novels:
        return
    print(f"加载了 {len(novels)} 篇小说")
    build_xlsx(novels, OUTPUT_XLSX)


if __name__ == "__main__":
    main()
